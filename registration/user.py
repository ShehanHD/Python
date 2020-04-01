from os.path import exists
import self as self
from openpyxl import Workbook, load_workbook
import datetime
import os
from app import *


class User:
    def __init__(self, arr):
        self.user = []
        exists = os.path.isfile('Users.xlsx')

        if exists is False:
            self.create_file(arr)
        else:
            self.read_file()

    def create_file(self, arr):
        # print(exists)
        j = 2
        wb = Workbook()  # crea oggetto
        sheet = wb.active  # attivatre il sheet

        sheet['A1'] = "Date"

        for i in arr:
            if j <= (len(arr)):
                sheet.cell(1, j, i)
            else:
                pass
            j += 1

        wb.save('Users.xlsx')

    def read_file(self):
        fp = load_workbook('Users.xlsx')
        sheet = fp['Sheet']

        for i in range(2, (sheet.max_row + 1)):
            a = []
            for j in range(1, sheet.max_column + 1):
                cell = sheet.cell(i, j)
                a.append(cell.value)
            self.user.append(a)

        fp.save("Users.xlsx")


class Register(User):
    def add_user(self, x):
        a = []
        password = False

        a.append(datetime.datetime.now())
        for i in range(len(x)):

            if i < (len(x)-2):
                a.append(input(f"{x[i]}> "))

            elif i == (len(x)-2):

                while password != True:
                    pwd1 = input(f"{x[len(x)-2]} > ")
                    pwd2 = input(f"{x[len(x)-1]} > ")
                    password = self.pwd_control(pwd1, pwd2)

                    if not password:
                        print("Password non sono uguali!")
                    else:
                        print("Inserimento è andato in buon fine!")
                        a.append(pwd1)
                        fp = load_workbook('Users.xlsx')
                        sheet = fp['Sheet']
                        for row in range((sheet.max_row + 1), (sheet.max_row + 2)):
                            for col in range(1, (sheet.max_column + 1)):
                                cell = sheet.cell(row, col)
                                cell.value = a[col-1]
                fp.save('Users.xlsx')

            else:
                pass

        self.user.append(a)
        print(self.user)

    def pwd_control(self, pwd, pwd2):
        return True if pwd == pwd2 else False


class Login(User):
    def control_login(self):
        print(self.user)