import openpyxl as xl

fp = xl.load_workbook('transactions.xlsx')

#print(fp)
sheet = fp['Sheet1']


for i in range(2, (sheet.max_row+1)):
    cell = sheet.cell(i, 3)         #sheet['A1']
    new = cell.value * 10
    new_col = sheet.cell(i, 4)
    new_col.value = new

fp.save('test2.xlsx')
